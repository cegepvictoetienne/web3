import argparse

def generer_grille(fichier, nom_feuille):
    import openpyxl

    excel = openpyxl.load_workbook(fichier, data_only=True)

    feuille = excel[nom_feuille]

    padding_de_grille = '|&#8288 {: style="padding:0"}'

    grille = []

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    grille.append(entete)
    grille.append(trait)

    for row in range(2, feuille.max_row + 1):
        ligne = ""
        has_colspan = False
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur).replace("\n", "")
            ligne += valeur
            if col < feuille.max_column:
                if not has_colspan:
                    ligne += "|"
            else:
                if has_colspan:
                    ligne += padding_de_grille * (feuille.max_column - 1)
                ligne += "\n"
            if "colspan" in valeur:
                has_colspan = True

        

        grille.append(ligne)
    
    return grille

def generer_horaire(titre_page=None, chemin_excel=None, chemin_markdown=None):
    if titre_page is None or chemin_excel is None:
        parser = argparse.ArgumentParser(
            description="Génère la page d'horaire à partir d'un fichier Excel."
        )
        parser.add_argument("--titre_page", help="Titre de la page Markdown")
        parser.add_argument("--chemin_excel", help="Chemin du fichier Excel source")
        parser.add_argument(
            "--chemin_markdown",
            default="./docs/horaire.md",
            help="Chemin du fichier Markdown de sortie (par défaut: ./docs/horaire.md)",
        )
        args = parser.parse_args()
        titre_page = args.titre_page
        chemin_excel = args.chemin_excel
        chemin_markdown = args.chemin_markdown

    grille = generer_grille(chemin_excel, "horaire")

    with open(chemin_markdown, "w", encoding="utf-8") as f:
        f.write(f"# {titre_page}\n")
        f.writelines(grille)
        
print("Générer la page de l'horaire")
generer_horaire()